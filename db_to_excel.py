# coding=utf-8
"""
db_to_excel.py
==============
功能：將 output_db 資料夾下的 .db 檔案轉換為 Excel 報表。
邏輯：
  1. 每個 DB 獨立轉換。
  2. 自動去重：同車牌 5 分鐘內重複，保留次數較高者。
  3. 自動分卷：超過 5000 筆自動拆分。
  4. 多進程預處理圖片：PIL 縮放 + JPEG 壓縮平行處理，大幅加速。
"""

import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PilImage
import io
import sys
import math
import multiprocessing as mp
from functools import partial
from pathlib import Path
from datetime import datetime, timedelta
from tqdm import tqdm

# ==========================================
# 1. 系統配置區
# ==========================================
DB_DIR = Path("output_db")          # 輸入：存放 .db 檔案的資料夾
OUTPUT_DIR = Path("output_xlsx")    # 輸出：產生 .xlsx 檔案的資料夾
LIMIT_PER_FILE = 5000               # 每份 Excel 最大筆數，超過自動分卷
TIME_THRESHOLD_MINUTES = 5          # 去重時間閾值（分鐘）：同車牌在此時間內視為重複
BATCH_SIZE = 200                    # 每個 worker 一次處理的圖片數量


# ==========================================
# 2. 圖片處理函式
# ==========================================
def resize_image_for_excel(image_path, target_width=120):
    """
    讀取單張圖片並縮放為指定寬度，壓縮為 JPEG bytes。
    
    :param image_path: 圖片檔案路徑
    :param target_width: 縮放後的目標寬度 (px)
    :return: (jpeg_bytes, 縮放後高度) 或 (None, 0) 表示失敗
    """
    path = Path(image_path)
    if not path.exists():
        return None, 0

    try:
        with PilImage.open(path) as img:
            # 依據目標寬度等比例縮放
            ratio = target_width / float(img.size[0])
            new_h = int(img.size[1] * ratio)
            img = img.resize((target_width, new_h), PilImage.LANCZOS).convert("RGB")

            # 壓縮為 JPEG 格式的 bytes（不寫入磁碟）
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=70)
            buf.seek(0)
            return buf.getvalue(), new_h  # 回傳 bytes（可跨進程序列化）
    except Exception:
        return None, 0


def process_image_batch(batch):
    """
    批次處理多張圖片的縮放與壓縮（由子進程呼叫）。
    每個 worker 一次處理一批，減少進程間通訊開銷。

    :param batch: [(df_idx, 車輛截圖路徑, 車牌截圖路徑), ...]
    :return: {df_idx: {'veh': (bytes, h), 'plate': (bytes, h)}, ...}
    """
    results = {}
    for idx, veh_path, plate_path in batch:
        # 分別處理車輛截圖與車牌截圖
        veh_data = resize_image_for_excel(veh_path) if veh_path else (None, 0)
        plate_data = resize_image_for_excel(plate_path) if plate_path else (None, 0)
        results[idx] = {'veh': veh_data, 'plate': plate_data}
    return results


def preprocess_images_parallel(df, pbar=None):
    """
    多進程平行預處理 DataFrame 中所有圖片。
    這是整個流程最耗時的步驟，透過多核心平行大幅加速。

    :param df: 去重後的資料 DataFrame（需含 veh_img_path, plate_img_path 欄位）
    :param pbar: 外部傳入的 tqdm 進度條，每處理一批就更新對應筆數
    :return: {df_idx: {'veh': (bytes, h), 'plate': (bytes, h)}, ...} 圖片快取字典
    """
    # 從 DataFrame 收集所有需要處理的圖片路徑
    tasks = []
    for idx, row in df.iterrows():
        tasks.append((idx, row.get('veh_img_path', ''), row.get('plate_img_path', '')))

    # 將任務切分成多個批次，每批 BATCH_SIZE 筆
    batches = [tasks[i:i + BATCH_SIZE] for i in range(0, len(tasks), BATCH_SIZE)]

    # worker 數量不超過 CPU 核心數，也不超過批次數
    num_workers = min(mp.cpu_count(), len(batches))
    all_results = {}

    with mp.Pool(processes=num_workers) as pool:
        # imap_unordered：不保證順序但更快，結果透過 idx 對應回 DataFrame
        for batch_result in pool.imap_unordered(process_image_batch, batches):
            all_results.update(batch_result)
            # 更新外部進度條（以筆數為單位）
            if pbar is not None:
                pbar.update(len(batch_result))

    return all_results


# ==========================================
# 3. 去重邏輯
# ==========================================
def filter_duplicates(df, db_name=""):
    """
    針對 DataFrame 進行去重。
    規則：同車號 (plate_text) 若時間差 < TIME_THRESHOLD_MINUTES 分鐘，
    保留 plate_count 較大者（次數多的較可靠）。

    :param df: 原始 DataFrame
    :param db_name: DB 名稱（用於 log 顯示）
    :return: 去重後的 DataFrame
    """
    if df.empty:
        return df

    before_count = len(df)

    # 將 video_time (HH:MM:SS) 轉為 datetime 物件以便計算時間差
    try:
        df['temp_dt'] = pd.to_datetime(df['video_time'], format='%H:%M:%S')
    except Exception as e:
        tqdm.write(f"    [{db_name}] 時間格式解析失敗，跳過去重 ({e})")
        return df

    # 依時間排序，確保後續逐筆比對的順序正確
    df = df.sort_values(by=['temp_dt']).reset_index(drop=True)

    indices_to_drop = []       # 記錄要刪除的列索引
    active_records = {}        # 每個車牌「目前保留」的最佳紀錄
    threshold = timedelta(minutes=TIME_THRESHOLD_MINUTES)

    for idx, row in df.iterrows():
        plate = row['plate_text']
        curr_time = row['temp_dt']
        curr_count = row['plate_count']

        # 新車牌：直接記錄為該車牌的保留紀錄
        if plate not in active_records:
            active_records[plate] = {'idx': idx, 'time': curr_time, 'cnt': curr_count}
            continue

        last_rec = active_records[plate]
        time_diff = curr_time - last_rec['time']

        if abs(time_diff) <= threshold:
            # 時間差在閾值內 → 視為重複，保留次數較高者
            if curr_count >= last_rec['cnt']:
                # 當前這筆較好（或相等取後者），刪除舊的
                indices_to_drop.append(last_rec['idx'])
                active_records[plate] = {'idx': idx, 'time': curr_time, 'cnt': curr_count}
            else:
                # 舊的較好，刪除當前這筆
                indices_to_drop.append(idx)
        else:
            # 時間差超過閾值 → 視為新的獨立事件
            active_records[plate] = {'idx': idx, 'time': curr_time, 'cnt': curr_count}

    # 刪除標記的重複列，移除暫存欄位，恢復原始排序
    df_clean = df.drop(indices_to_drop)
    df_clean = df_clean.drop(columns=['temp_dt']).sort_values(by='id').reset_index(drop=True)

    tqdm.write(f"  [{db_name}] 去重: {before_count} → {len(df_clean)} 筆 (刪除 {len(indices_to_drop)} 筆)")
    return df_clean


# ==========================================
# 4. Excel 寫入
# ==========================================
def write_dataframe_to_excel(df, output_path, image_cache, pbar=None):
    """
    將 DataFrame 寫入 Excel 檔案，圖片從預處理快取中取用。
    openpyxl 不是 thread-safe，因此寫入必須在單一執行緒中進行。
    但因為圖片已經是壓縮好的 bytes，插入速度非常快。

    :param df: 要寫入的資料 DataFrame
    :param output_path: 輸出的 .xlsx 檔案路徑
    :param image_cache: 圖片快取字典 {idx: {'veh': (bytes, h), 'plate': (bytes, h)}}
    :param pbar: 外部傳入的 tqdm 進度條（可選）
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "車輛辨識結果"

    # 寫入標頭列
    headers = ["ID", "車輛截圖", "車種", "車牌截圖", "車號", "次數", "時間軸", "時間點"]
    ws.append(headers)

    # 設定各欄位寬度（讓圖片和文字顯示合理）
    ws.column_dimensions["B"].width = 18   # 車輛截圖欄
    ws.column_dimensions["D"].width = 18   # 車牌截圖欄
    ws.column_dimensions["G"].width = 15   # 時間軸欄
    ws.column_dimensions["H"].width = 20   # 時間點欄

    # 逐筆寫入資料與圖片
    for idx, row in df.iterrows():
        excel_row = idx + 2   # Excel 第 1 列是標頭，資料從第 2 列開始

        # 寫入文字欄位
        ws.cell(excel_row, 1, row['track_id'])
        ws.cell(excel_row, 3, row['class_name'])
        ws.cell(excel_row, 5, row['plate_text'])
        ws.cell(excel_row, 6, row['plate_count'])
        ws.cell(excel_row, 7, row['video_time'])
        ws.cell(excel_row, 8, row['real_time'])

        # 從快取取出預處理好的圖片 bytes
        cached = image_cache.get(idx, {})

        # 插入車輛截圖（B 欄）
        veh_bytes, veh_h = cached.get('veh', (None, 0))
        if veh_bytes:
            img = XLImage(io.BytesIO(veh_bytes))
            img.anchor = f"B{excel_row}"
            ws.add_image(img)
            ws.row_dimensions[excel_row].height = veh_h * 0.75  # 像素轉換為 Excel 列高

        # 插入車牌截圖（D 欄）
        plate_bytes, plate_h = cached.get('plate', (None, 0))
        if plate_bytes:
            img = XLImage(io.BytesIO(plate_bytes))
            img.anchor = f"D{excel_row}"
            ws.add_image(img)
            # 取兩張圖片中較高的那個作為列高
            current_h = ws.row_dimensions[excel_row].height or 0
            if plate_h * 0.75 > current_h:
                ws.row_dimensions[excel_row].height = plate_h * 0.75

        # 更新外部進度條
        if pbar is not None:
            pbar.update(1)

    # 儲存 Excel 檔案
    wb.save(output_path)


# ==========================================
# 5. 單一 DB 處理流程
# ==========================================
def process_single_db(db_path, pbar):
    """
    處理單一 .db 檔案的完整流程：
    讀取 DB → 去重 → 多進程預處理圖片 → 寫入 Excel（必要時自動分卷）

    :param db_path: .db 檔案的 Path 物件
    :param pbar: 外部總進度條（以筆數為單位）
    """
    # 讀取資料庫
    try:
        conn = sqlite3.connect(db_path)
        df = pd.read_sql_query("SELECT * FROM vehicle_records ORDER BY id", conn)
        conn.close()
    except Exception as e:
        tqdm.write(f"\n[錯誤] 無法讀取 DB {db_path}: {e}")
        return

    if df.empty:
        return

    # 步驟 1：去重（同車牌 5 分鐘內取次數最高者）
    df = filter_duplicates(df, db_name=db_path.stem)
    total_records = len(df)

    # 步驟 2：多進程平行預處理所有圖片（最耗時的部分）
    # 進度條描述切換為「預處理圖片」，讓使用者知道目前在做什麼
    pbar.set_description(f"{db_path.stem} | 預處理圖片")
    image_cache = preprocess_images_parallel(df, pbar)

    # 步驟 3：寫入 Excel（圖片已是 bytes 快取，插入很快）
    base_name = db_path.stem

    if total_records <= LIMIT_PER_FILE:
        # 不需分卷：直接寫成一份 Excel
        out_path = OUTPUT_DIR / f"{base_name}_output.xlsx"
        pbar.set_description(f"{db_path.stem} | 寫入 Excel")
        write_dataframe_to_excel(df, out_path, image_cache)
    else:
        # 需要分卷：依照 LIMIT_PER_FILE 切分
        total_parts = math.ceil(total_records / LIMIT_PER_FILE)

        for i in range(total_parts):
            start_idx = i * LIMIT_PER_FILE
            end_idx = min((i + 1) * LIMIT_PER_FILE, total_records)

            # 切出該分卷的資料子集，重設索引從 0 開始
            df_part = df.iloc[start_idx:end_idx].reset_index(drop=True)

            # 重新映射 image_cache 的 key（因為 reset_index 後 idx 從 0 開始）
            part_cache = {}
            for new_idx, old_idx in enumerate(range(start_idx, end_idx)):
                if old_idx in image_cache:
                    part_cache[new_idx] = image_cache[old_idx]

            out_path = OUTPUT_DIR / f"{base_name}_output_part{i + 1}.xlsx"
            pbar.set_description(f"{db_path.stem} | 寫入 Part {i + 1}/{total_parts}")
            write_dataframe_to_excel(df_part, out_path, part_cache)


# ==========================================
# 6. 主程式進入點
# ==========================================
def main():
    """
    主函式：掃描所有 .db 檔案，依序處理。
    先快速計算總筆數，建立單一 tqdm 進度條，
    動態顯示當前 DB 名稱與處理階段。
    """
    if not DB_DIR.exists():
        print(f"[錯誤] 找不到資料夾: {DB_DIR}")
        return

    db_files = list(DB_DIR.glob("*.db"))
    if not db_files:
        print(f"[提示] 在 {DB_DIR} 找不到任何 .db 檔案。")
        return

    # 預先掃描所有 DB 的筆數（僅 SELECT COUNT，非常快）
    total_all = 0
    for db_path in db_files:
        try:
            conn = sqlite3.connect(db_path)
            count = pd.read_sql_query(
                "SELECT COUNT(*) as cnt FROM vehicle_records", conn
            ).iloc[0]['cnt']
            conn.close()
            total_all += count
        except Exception:
            pass

    print(f"[INFO] {len(db_files)} 個資料庫，共約 {total_all} 筆")

    # 建立單一總進度條，以「筆」為單位貫穿全程
    with tqdm(total=total_all, unit="筆", dynamic_ncols=True) as pbar:
        for db_path in db_files:
            process_single_db(db_path, pbar)

    print(f"✅ 全部完成！請查看 {OUTPUT_DIR} 資料夾。")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n🛑 中斷")
        sys.exit(0)