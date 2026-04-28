# coding=utf-8
"""
excel_utils.py
輸出 Excel（包含圖片縮圖）
"""

import io
from pathlib import Path
from PIL import Image as PilImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

def resize_to_fixed_width(image_path, target_width=150, quality=70):
    # 檔案不存在就跳過
    if not image_path or not Path(image_path).exists():
        return b"", (0, 0)

    with PilImage.open(image_path) as img:
        # 維持比例縮放
        ratio = target_width / float(img.size[0])
        new_h = int(img.size[1] * ratio)

        img = img.resize((target_width, new_h), PilImage.LANCZOS).convert("RGB")

        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=quality, optimize=True)
        buf.seek(0)
        return buf.getvalue(), (target_width, new_h)

def save_to_excel(stats_list, excel_path):
    if not stats_list:
        print("[INFO] 無資料可寫入 Excel")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "車輛統計"

    # header
    ws.append(["ID", "車輛截圖", "車種", "車牌截圖", "車號", "次數", "時間軸", "時間點", "執行時間"])
    
    # 設定欄寬
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["D"].width = 15

    for r, d in enumerate(stats_list, start=2):
        # 寫入文字資料
        ws.cell(r, 1, d["ID"])
        ws.cell(r, 2, "")
        ws.cell(r, 3, d["車種"])
        ws.cell(r, 4, "")
        ws.cell(r, 5, d["車號"])
        ws.cell(r, 6, d["次數"])
        ws.cell(r, 7, d["時間軸"])
        ws.cell(r, 8, d["時間點"])
        ws.cell(r, 9, d["電腦執行時間"])

        # 寫入車身圖片
        vb, (vw, vh) = resize_to_fixed_width(d["veh_path"], 100)
        if vb:
            im = XLImage(io.BytesIO(vb))
            im.anchor = f"B{r}"
            ws.add_image(im)

        # 寫入車牌圖片
        pb, (pw, ph) = resize_to_fixed_width(d["lpr_path"], 100)
        if pb:
            im = XLImage(io.BytesIO(pb))
            im.anchor = f"D{r}"
            ws.add_image(im)

        # 設定列高
        ws.row_dimensions[r].height = max(vh, ph) * 0.75

    # ⭐⭐ [關鍵修正] 確保目錄存在，避免 FileNotFoundError ⭐⭐
    Path(excel_path).parent.mkdir(parents=True, exist_ok=True)
    
    # 存檔
    wb.save(excel_path)
    # print(f"[INFO] 統計結果已更新至 {excel_path}")