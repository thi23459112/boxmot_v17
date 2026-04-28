# coding=utf-8
"""
state_excel_track.py
==============
純車流統計專用狀態管理（批次寫入版）：
- 車輛 IN/OUT 方向判定
- 根據 y_first 與 y_last 的 Y 軸位移判定進出方向
- 達到結算門檻後批次寫入統計，避免每筆重寫 Excel 導致卡頓
"""

import io
from collections import Counter, deque
from pathlib import Path
import datetime
import time
import cv2
import threading
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage

from . import config
from .vision_utils import class_name
from . import excel_utils

# ============================================================
# 批次寫入控制（與 state_excel.py 相同邏輯）
# ============================================================
_global_stats_list = deque(maxlen=10000)  # 最多保留 1 萬筆
_lock = threading.Lock()

_BATCH_SIZE = 100                    # 累積 100 筆就寫入
_FLUSH_INTERVAL_SECONDS = 300        # 或超過 300 秒強制寫入
_last_flush_time = time.time()
_flush_lock = threading.Lock()

def _get_excel_path():
    """取得 Excel 輸出路徑 (優先讀 config)"""
    p = getattr(config, "RESULT_EXCEL_PATH", None)
    if p:
        return Path(p)
    return Path("output_xlsx") / "車流統計_完整版.xlsx"

def _save_now():
    """將目前的列表批次寫入 Excel（追加模式，不重寫歷史資料）"""
    global _global_stats_list
    
    path = _get_excel_path()
    
    with _lock:
        data_to_write = list(_global_stats_list)
        _global_stats_list.clear()
    
    if data_to_write:
        _append_to_excel(data_to_write, str(path))
        print(f"[INFO] 批次寫入 {len(data_to_write)} 筆車流資料到 Excel")

def _append_to_excel(new_data, excel_path):
    """
    追加寫入 Excel：讀取現有檔案，在最後面加入新資料
    避免每次重寫所有歷史資料
    """
    Path(excel_path).parent.mkdir(parents=True, exist_ok=True)
    
    # 如果檔案存在，讀取現有工作表
    if Path(excel_path).exists():
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            start_row = ws.max_row + 1
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.title = "車流統計"
            ws.append(["ID", "電腦執行時間", "時間軸", "時間點", "車種", "次數", "區域"])
            start_row = 2
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "車流統計"
        ws.append(["ID", "電腦執行時間", "時間軸", "時間點", "車種", "次數", "區域"])
        start_row = 2
    
    # 寫入新資料
    for idx, d in enumerate(new_data):
        r = start_row + idx
        ws.cell(r, 1, d["ID"])
        ws.cell(r, 2, d["電腦執行時間"])
        ws.cell(r, 3, d["時間軸"])
        ws.cell(r, 4, d["時間點"])
        ws.cell(r, 5, d["車種"])
        ws.cell(r, 6, d["次數"])
        ws.cell(r, 7, d["區域"])
    
    wb.save(excel_path)

def _maybe_flush(force=False):
    """檢查是否需要批次寫入"""
    global _last_flush_time
    
    with _flush_lock:
        now = time.time()
        should_flush = (
            force or 
            len(_global_stats_list) >= _BATCH_SIZE or
            now - _last_flush_time >= _FLUSH_INTERVAL_SECONDS
        )
        
        if should_flush and len(_global_stats_list) > 0:
            _save_now()
            _last_flush_time = now

def flush_all_sessions_now(reason="manual"):
    """
    提供給 main_track.py 的介面 (按 Q 時呼叫)
    強制寫入所有剩餘資料
    """
    print(f"[INFO] 強制寫入車流資料 ({reason})，剩餘 {len(_global_stats_list)} 筆...")
    _maybe_flush(force=True)

# ============================================================
# State 邏輯（純車流）
# ============================================================

def init_state():
    """ 
    初始化車流專用狀態字典
    """
    return {
        "roi_recent": [],           # 存放 ROI 內的 class_id (投票用)
        "classes": Counter(),       # 類別計數
        "frames_since_left": 0,     # 離開 ROI 幀數
        "missed_frames": 0,         # 連續沒 seen 幀數
        "counted": False,           # 是否已結算
        "last_in_roi": False,       # 最後一次是否在 ROI 內
        "y_first": None,            # 第一次進入 ROI 的 Y 座標
        "y_last": None,             # 最後一次出現在 ROI 的 Y 座標
        "first_vsec": None,         # 第一次進入 ROI 的秒數（影片時間軸）
        "roi_hits": 0,              # 累計 ROI 命中次數
        "first_seen_frame": 0,      # 第一次看到的幀號（用於清理孤兒）
    }

def consider_cleanup_and_finalize(
    tid: int,
    st: dict,
    seen: bool,
    in_roi: bool,
    class_names,
    current_world_time: datetime.datetime,
    stats_list: list,  # 為相容性保留，但不再使用
    current_frame_id: int = 0,  # 當前幀號，用於清理孤兒
):
    """
    純車流結算邏輯：
    1) 更新 missing 狀態（遮擋容忍）
    2) 更新離開計數與最後狀態
    3) 結算條件：離開 ROI 超過設定幀數且尚未結算
    4) 方向判定：根據 y_first 與 y_last 的 Y 軸位移判定 IN/OUT
    5) 清除條件：結算後離開太久，從記憶體刪除
    """

    # 0) 紀錄第一次看到的幀號（用於孤兒清理）
    if st.get("first_seen_frame") == 0:
        st["first_seen_frame"] = current_frame_id

    # 1) 更新 Missing 狀態
    if not seen:
        st["missed_frames"] += 1

    # 2) 更新離開計數與最後狀態
    if in_roi:
        st["frames_since_left"] = 0
        st["roi_hits"] += 1
    elif seen:
        st["frames_since_left"] += 1
    else:
        if st["last_in_roi"] and st["missed_frames"] <= config.MISS_GRACE_FRAMES:
            st["frames_since_left"] = 0
        else:
            st["frames_since_left"] += 1

    st["last_in_roi"] = in_roi

    # 3) 結算條件：離開 ROI 超過設定幀數且尚未結算
    if st["frames_since_left"] > config.LEAVE_ROI_FRAMES_TO_COUNT and not st["counted"]:
        
        # 方向判定邏輯
        direction = "NA"
        if st["y_first"] is not None and st["y_last"] is not None:
            diff_y = st["y_last"] - st["y_first"]
            
            if diff_y > config.MOVEMENT_THRESHOLD_PX: 
                direction = "IN"
            elif diff_y < -config.MOVEMENT_THRESHOLD_PX: 
                direction = "OUT"

        # 判定是否符合寫入統計的標準
        if direction != "NA" and st["roi_hits"] >= config.MIN_ROI_HITS:
            if st["roi_recent"]:
                cls_major, cnt = Counter(st["roi_recent"]).most_common(1)[0]
                cls_txt = class_name(class_names, cls_major)
                
                # ⭐ 修正：時間軸使用影片時間（秒），時間點使用真實世界時間
                video_sec = st.get("first_vsec", 0.0) or 0.0
                time_axis = time.strftime("%H:%M:%S", time.gmtime(video_sec))
                time_point = current_world_time.strftime("%Y-%m-%d %H:%M:%S")

                row_data = {
                    "ID": tid,
                    "電腦執行時間": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "時間軸": time_axis,  # 影片時間軸 00:00:12
                    "時間點": time_point,  # 真實時間 2025-05-22 08:30:15
                    "車種": cls_txt,
                    "次數": cnt,
                    "區域": direction
                }
                
                # ⭐ 批次寫入，而非直接 append 到 stats_list
                with _lock:
                    _global_stats_list.append(row_data)
                
                _maybe_flush(force=False)
                
                print(f"[統計結算] ID={tid}, 方向={direction}, 車種={cls_txt}, ROI命中={st['roi_hits']}, 影片時間軸={time_axis}")

        st["counted"] = True

    # 4) 清除條件：結算後離開太久，或孤兒清理
    should_remove = False
    
    # 正常清理：已結算且離開夠久
    if st["counted"] and st["frames_since_left"] > config.CLEANUP_FRAMES:
        should_remove = True
    
    # 孤兒清理：從未結算，但消失太久（超過 300 幀）
    ORPHAN_MAX_AGE = 300
    if not st["counted"] and st["missed_frames"] > ORPHAN_MAX_AGE:
        should_remove = True

    return should_remove