# coding=utf-8
"""
state_excel.py
==============
記憶體優化版：解決長時運行導致的 OOM (Out Of Memory)
- 策略：小量批次寫入 (Batch Flush) + 強制 GC
- 邏輯：
    1. 每累積 200 筆 (FLUSH_THRESHOLD) -> 寫入硬碟 (Append 模式) -> 釋放 RAM
    2. 每寫滿 5000 筆 (ENTRIES_PER_FILE) -> 換下一個 Excel 檔案
"""

import io
from collections import Counter, deque
from pathlib import Path
import datetime
import time
import cv2
import threading
import queue
import gc  # ⭐ 新增：垃圾回收模組

# 引入 excel_utils
from . import excel_utils
from .vision_utils import class_name
from . import config

# ============================================================
# 記憶體控制參數
# ============================================================
# ⭐ 關鍵修改：記憶體內最多只留 200 筆，滿了就寫入硬碟
_FLUSH_THRESHOLD = 200                      # 批次寫入的觸發門檻 (累積筆數)

# ⭐ 檔案控制：雖然每 200 筆寫一次，但每 5000 筆才換一個檔案 (避免檔案碎片化)
_ENTRIES_PER_FILE = 5000                    # 每個 Excel 檔案的最大記錄筆數

# 全域資料容器：存放尚未寫入的統計資料 (批次緩衝區)
_global_stats_list = []                     # 改用 list，因為我們會頻繁清空
_lock = threading.Lock()                    # 保護 _global_stats_list 的執行緒鎖

# 背景寫入相關
_write_queue = queue.Queue(maxsize=10)      # 佇列不用太大，避免積壓過多記憶體
_writer_thread = None                       # 背景寫入執行緒物件
_writer_running = False                     # 控制背景執行緒運行的旗標

# 檔案段號計數：用於分段儲存 Excel 檔案 (如 第1段、第2段...)
_current_file_index = 1
_entries_in_current_file = 0                # 當前檔案已寫入的筆數

# 時間閥值 (防止資料量太少一直不寫入)
_FLUSH_INTERVAL_SECONDS = 60                # 最大等待時間 (秒)，超過則強制寫入
_last_flush_time = time.time()              # 上次寫入的時間戳記
_flush_lock = threading.Lock()              # 保護時間相關變數 (若需要可擴充)


def _get_excel_path(segment=None):
    """取得 Excel 輸出路徑 (分段命名)
    :param segment: 檔案段號，若為 None 則回傳基礎路徑 (不帶段號)
    :return: Path 物件
    """
    p = getattr(config, "RESULT_EXCEL_PATH", None)
    if p:
        base_path = Path(p)
    else:
        base_path = Path("output_xlsx") / "車輛統計_完整版.xlsx"
    
    if segment is not None:
        stem = base_path.stem           # 主檔名 (不含副檔名)
        suffix = base_path.suffix       # 副檔名 (如 .xlsx)
        return base_path.parent / f"{stem}_第{segment}段{suffix}"
    
    return base_path


def _background_writer():
    """
    背景執行緒：持續從佇列取資料 -> 寫入 Excel -> 強制 GC
    """
    global _writer_running, _current_file_index, _entries_in_current_file
    
    print("[INFO] Excel 背景寫入執行緒啟動 (記憶體優化模式)")
    
    while _writer_running:
        try:
            # 從佇列取得一批資料 (等待最多 1 秒)
            data_to_write = _write_queue.get(timeout=1.0)
            
            if data_to_write is None:    # 結束訊號
                break
            
            count = len(data_to_write)   # 本批資料筆數
            
            # 判斷是否需要換檔 (若加上這批會超過上限)
            if _entries_in_current_file + count > _ENTRIES_PER_FILE:
                # 如果這批寫進去會爆，先寫一部分填滿當前檔案 (這裡簡化處理：直接換新檔)
                # 為了避免邏輯太複雜，這裡選擇直接換新檔
                _current_file_index += 1
                _entries_in_current_file = 0
                print(f"[INFO] 檔案已滿，切換至：第{_current_file_index}段")
            
            # 執行寫入
            path = _get_excel_path(_current_file_index)
            _do_write_to_excel(data_to_write, str(path))
            
            # 更新計數
            _entries_in_current_file += count
            
            # ⭐⭐ 關鍵優化：寫完後強制清理記憶體 ⭐⭐
            del data_to_write            # 刪除本批資料參照
            gc.collect()                 # 手動觸發垃圾回收
            
        except queue.Empty:
            continue                     # 佇列空閒，繼續迴圈
        except Exception as e:
            print(f"[ERROR] 背景寫入失敗: {e}")
            import traceback
            traceback.print_exc()

    print("[INFO] Excel 背景寫入執行緒結束")


def _start_writer_thread():
    """啟動背景寫入執行緒 (若尚未啟動)"""
    global _writer_thread, _writer_running
    if _writer_thread is None or not _writer_thread.is_alive():
        _writer_running = True
        _writer_thread = threading.Thread(target=_background_writer, name="ExcelWriter", daemon=True)
        _writer_thread.start()


def _stop_writer_thread():
    """停止背景寫入執行緒 (等待佇列清空)"""
    global _writer_running, _writer_thread
    _writer_running = False
    if _writer_thread is not None:
        try:
            _write_queue.put(None, timeout=1.0)    # 放入結束訊號
        except queue.Full:
            pass
        _writer_thread.join(timeout=5.0)           # 等待執行緒結束
        _writer_thread = None


def _do_write_to_excel(new_data, excel_path):
    """
    Append 模式寫入：讀取現有檔案 -> 附加資料 -> 存檔
    :param new_data: 要寫入的新資料列表 (每個元素為 dict)
    :param excel_path: Excel 檔案路徑
    """
    from openpyxl import load_workbook
    
    Path(excel_path).parent.mkdir(parents=True, exist_ok=True)
    
    # 嘗試讀取現有檔案
    wb = None
    try:
        if Path(excel_path).exists():
            wb = load_workbook(excel_path)
            ws = wb.active
            start_row = ws.max_row + 1                  # 從下一列開始追加
        else:
            wb = excel_utils.Workbook()                 # 建立新活頁簿
            ws = wb.active
            ws.title = "車輛統計"
            # 寫入標題列
            ws.append(["ID", "車輛截圖", "車種", "車牌截圖", "車號", "次數", "時間軸", "時間點", "執行時間"])
            ws.column_dimensions["B"].width = 15        # 設定車輛截圖欄位寬度
            ws.column_dimensions["D"].width = 15        # 設定車牌截圖欄位寬度
            start_row = 2                               # 資料從第二列開始
        
        # 寫入資料
        for idx, d in enumerate(new_data):
            r = start_row + idx
            
            # 填入基本資料 (不含圖片)
            ws.cell(r, 1, d["ID"])
            ws.cell(r, 2, "")            # 車輛截圖欄位留空 (待會插入圖片)
            ws.cell(r, 3, d["車種"])
            ws.cell(r, 4, "")            # 車牌截圖欄位留空
            ws.cell(r, 5, d["車號"])
            ws.cell(r, 6, d["次數"])
            ws.cell(r, 7, d["時間軸"])
            ws.cell(r, 8, d["時間點"])
            ws.cell(r, 9, d["電腦執行時間"])

            # 處理圖片 (這是最吃記憶體的部分)
            try:
                if d.get("veh_path"):
                    # 將車輛圖片縮放至固定寬度 100，並取得二進位資料
                    vb, (vw, vh) = excel_utils.resize_to_fixed_width(d["veh_path"], 100)
                    if vb:
                        im = excel_utils.XLImage(io.BytesIO(vb))   # 建立 Excel 圖片物件
                        im.anchor = f"B{r}"                        # 錨定在 B 欄該列
                        ws.add_image(im)

                if d.get("lpr_path"):
                    pb, (pw, ph) = excel_utils.resize_to_fixed_width(d["lpr_path"], 100)
                    if pb:
                        im = excel_utils.XLImage(io.BytesIO(pb))
                        im.anchor = f"D{r}"
                        ws.add_image(im)
                
                # 設定列高：取車輛圖片和車牌圖片高度的最大值，轉換為 Excel 單位
                h_val = max(vh, ph) if 'vh' in locals() and 'ph' in locals() else 75
                ws.row_dimensions[r].height = h_val * 0.75
                
            except Exception as e:
                print(f"[WARNING] 圖片處理失敗 ID={d['ID']}: {e}")

        wb.save(excel_path) # 儲存檔案
        
    except Exception as e:
        print(f"[ERROR] Excel 存檔失敗: {e}")
    finally:
        # 確保 Workbook 關閉並釋放資源
        if wb:
            wb.close()
            del wb


def _trigger_flush():
    """
    將目前的 buffer 移交給背景執行緒
    """
    global _global_stats_list, _last_flush_time
    
    # 快速交換數據，避免鎖住太久
    with _lock:
        if not _global_stats_list:
            return
        data_to_write = _global_stats_list[:]    # 淺拷貝目前緩衝區
        _global_stats_list.clear()               # 清空全域 buffer
    
    _start_writer_thread()                       # 確保背景執行緒已啟動
    
    try:
        # 將數據放入佇列 (超時 0.1 秒)
        _write_queue.put(data_to_write, timeout=0.1)
        _last_flush_time = time.time()           # 更新最後寫入時間
        # print(f"[INFO] 已觸發寫入，數量: {len(data_to_write)}")
    except queue.Full:
        print("[WARNING] 寫入速度過慢，佇列已滿 (建議檢查硬碟IO)")


def _maybe_flush(force=False):
    """檢查是否滿足寫入條件 (筆數或時間)"""
    global _last_flush_time
    
    # 不需加鎖檢查長度，讀取是原子操作
    current_len = len(_global_stats_list)
    now = time.time()
    
    should_flush = False
    if force and current_len > 0:
        should_flush = True
    elif current_len >= _FLUSH_THRESHOLD: # 滿 200 筆
        should_flush = True
    elif current_len > 0 and (now - _last_flush_time >= _FLUSH_INTERVAL_SECONDS):  # 超過 60 秒
        should_flush = True
        
    if should_flush:
        _trigger_flush()


def flush_all_sessions_now(reason="manual"):
    """強制寫入 (程式結束用)"""
    print(f"[INFO] 強制寫入剩餘資料 ({reason})...")
    
    # 1. 先把 buffer 裡的丟進佇列
    _maybe_flush(force=True)
    
    # 2. 等待佇列清空 (停止背景執行緒)
    _stop_writer_thread()
    print("[INFO] 寫入完成")


# ============================================================
# State 邏輯 (介面維持不變)
# ============================================================

def init_state():
    """初始化一個新的狀態字典，用於追蹤單一車輛的資料"""
    return {
        "plates": Counter(),                    # 記錄此車輛出現過的車牌號碼及其次數
        "classes": Counter(),                   # 記錄此車輛被分類的類別及其次數
        "frames_since_left": 0,                 # 自上次離開 ROI 後的幀數
        "missed_frames": 0,                     # 追蹤器遺失此目標的連續幀數
        "counted": False,                       # 是否已統計過 (避免重複統計)
        "last_in_roi": False,                   # 上次幀是否在 ROI 內
        "plate_best": {},                       # 存放最佳車牌截圖的資訊 (鍵為車牌號碼，值為 dict)
        "enter_dt": None,                       # 此車輛首次進入 ROI 的日期時間
        "first_seen_frame": 0,                  # 此車輛首次出現的幀數編號
    }


def consider_cleanup_and_finalize(
    tid: int, st: dict, seen: bool, in_roi: bool, class_names,
    current_world_time: datetime.datetime, stats_list: list,
    current_frame_id: int = 0, video_time: float = 0.0,
):
    """
    考慮是否要結算此車輛的統計資料，並決定是否移除該狀態
    :param tid: 追蹤 ID
    :param st: 該車輛的狀態字典 (由 init_state 建立)
    :param seen: 當前幀是否出現此 ID
    :param in_roi: 當前幀是否在 ROI 內
    :param class_names: 類別編號對應的名稱字典
    :param current_world_time: 當前世界時間 (datetime)
    :param stats_list: 全域統計列表 (用於向後相容，但此處未使用)
    :param current_frame_id: 當前幀數編號
    :param video_time: 當前幀數在影片中的時間 (秒)
    :return: 布林值，True 表示該 ID 應從 states 中移除
    """
    now_dt = current_world_time

    # 記錄第一次出現的幀數編號
    if st.get("first_seen_frame") == 0:
        st["first_seen_frame"] = current_frame_id

    # 若目前位於 ROI 內且被看見，記錄進入時間
    if seen and in_roi:
        if st["enter_dt"] is None:
            st["enter_dt"] = now_dt
        st["last_in_roi"] = True
    elif not in_roi:
        st["last_in_roi"] = False

    # 更新遺失幀計數
    if not seen:
        st["missed_frames"] += 1
    else:
        st["missed_frames"] = 0

    # 更新離開 ROI 後的幀計數
    if in_roi:
        st["frames_since_left"] = 0
    elif seen:
        # 雖然不在 ROI 內但被看見，表示可能剛離開
        st["frames_since_left"] += 1
    else:
        # 未看見：若之前還在 ROI 內且遺失幀未超過寬限期，視為仍可能回來
        if st["last_in_roi"] and st["missed_frames"] <= config.MISS_GRACE_FRAMES:
            st["frames_since_left"] = 0
        else:
            st["frames_since_left"] += 1

    # 結算邏輯：當離開 ROI 的幀數超過門檻，且尚未統計過
    if st["frames_since_left"] > config.LEAVE_ROI_FRAMES_TO_COUNT and not st["counted"]:
        # 找出出現次數最多的車牌和類別
        most_p = st["plates"].most_common(1)
        most_c = st["classes"].most_common(1)

        # 需有車牌且票數足夠，且有類別
        if most_p and most_c and most_p[0][1] >= config.PLATE_MIN_VOTES:
            plate_str, plate_cnt = most_p[0]
            cls_id = int(most_c[0][0])
            cls_txt = class_name(class_names, cls_id)
            best = st["plate_best"].get(plate_str)

            if best is not None:
                video_sec = best.get("video_time", 0.0)                            # 車牌截圖對應的影片時間
                time_axis = time.strftime("%H:%M:%S", time.gmtime(video_sec))      # 格式化時間軸
                
                if st.get("enter_dt"):
                    real_time_str = st["enter_dt"].strftime("%Y-%m-%d %H:%M:%S")   # 真實進入時間
                else:
                    real_time_str = now_dt.strftime("%Y-%m-%d %H:%M:%S")           # 若無則用當前時間
                
                real_ts_file = now_dt.strftime("%m%d_%H%M%S")                      # 用於檔名時間戳

                # 建立儲存目錄
                v_dir = config.SCREENSHOT_DIR / str(cls_txt)
                p_dir = config.SCREENSHOT_DIR / "LPR"
                v_dir.mkdir(parents=True, exist_ok=True)
                p_dir.mkdir(parents=True, exist_ok=True)

                # 產生圖片儲存路徑
                v_path = v_dir / f"{tid}_{real_ts_file}.jpg"
                p_path = p_dir / f"{tid}_{plate_str}_{real_ts_file}.jpg"

                # 存圖 (IO 操作)
                cv2.imwrite(str(v_path), best["veh_img"])
                cv2.imwrite(str(p_path), best["plate_crop"])

                # 組合成一筆記錄
                row_data = {
                    "ID": tid,
                    "車輛截圖": str(v_path),
                    "車種": cls_txt,
                    "車牌截圖": str(p_path),
                    "車號": plate_str,
                    "次數": plate_cnt,
                    "時間軸": time_axis,
                    "時間點": real_time_str,
                    "電腦執行時間": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "veh_path": str(v_path),
                    "lpr_path": str(p_path)
                }

                print(f"[統計] ID={tid}, 車種={cls_txt}, 車號={plate_str}, 次數={plate_cnt}, 影片時間軸={time_axis}")
                
                # 將記錄加入全域緩衝區 (執行緒安全)
                with _lock:
                    _global_stats_list.append(row_data)
                
                # 檢查是否需要觸發寫入 (門檻 200 筆或時間超過 60 秒)
                _maybe_flush(force=False)

        st["counted"] = True
        st["plate_best"].clear()   # 重要：釋放圖片快取，避免記憶體洩漏

    # 決定是否應從 states 中移除該 ID
    should_remove = False
    # 已統計且離開 ROI 超過清除門檻
    if st["counted"] and st["frames_since_left"] > config.CLEANUP_FRAMES:
        should_remove = True
    
    # 孤兒清理：未被統計且遺失幀數超過最大存活時間 (300 幀)
    ORPHAN_MAX_AGE = 300
    if not st["counted"] and st["missed_frames"] > ORPHAN_MAX_AGE:
        should_remove = True

    return should_remove